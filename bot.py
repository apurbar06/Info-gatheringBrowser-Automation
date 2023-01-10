from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import time
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from tqdm import tqdm





# read parameters from file
with open('parameters.txt') as file:
    lines = [line.rstrip() for line in file]

_file_name = lines[0]
_sheet_name = lines[1]
_row = int(lines[2])

lines[0] = lines[0]+"\n"
lines[1] = lines[1]+"\n"



# read excel file
df = pd.ExcelFile(_file_name).parse(_sheet_name) 
name=df['Stock Name'].tolist()
number=df['BSE / NSE Code'].tolist()

chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
driver.maximize_window()
action = ActionChains(driver)
wait = WebDriverWait(driver, 5)
driver.get('https://www.morningstar.in')
time.sleep(10)
print()



for i in tqdm(range(0, len(name)), initial = _row-2, desc ="Progress: "):
    if(_row-2>=len(name)): 
        break
    print("\n\n\n\n"+str(name[_row-2])+" -> "+str(number[_row-2]))
    
    

    # Search and Enter
    # search_term = str(name[_row-2])
    search_number = str(number[_row-2])
    sbox = driver.find_element(by=By.XPATH, value='//input[@placeholder="Quote"]')
    sbox.send_keys(search_number)
    time.sleep(5)
    driver.find_element(by=By.XPATH, value="//div/strong[contains(text(),'"+search_number+"')]").click()
    # driver.find_element(by=By.XPATH, value="(.//div[@class='col-xs-12 stock-details']/strong)").click()
    time.sleep(10)


    # Extract rating
    rating = driver.find_element(by=By.XPATH, value="//div[contains(@class, 'mds-star-rating__sal mds-star-rating--large__sal')]").get_attribute("aria-label")
    rating = rating.split(" ")[2]
    # Extract the value of price/book
    price_pre_book = driver.find_element(by=By.XPATH, value="(.//ul[@class='small-block-grid-3 large-block-grid-4 sal-component-band-grid']/li)[10]//div[@class='dp-value']").text
    # Extract the value of price/cash flow
    price_pre_cash_flow = driver.find_element(by=By.XPATH, value="(.//ul[@class='sal-xsmall-block-grid-2 small-block-grid-4']/li)[2]//div[@class='dp-value']").text
    # Extract the value of price/sales
    price_pre_sales = driver.find_element(by=By.XPATH, value="(.//ul[@class='sal-xsmall-block-grid-2 small-block-grid-4']/li)[3]//div[@class='dp-value']").text
    # Extract the value of price/earnings
    price_pre_earnings = driver.find_element(by=By.XPATH, value="(.//ul[@class='sal-xsmall-block-grid-2 small-block-grid-4']/li)[4]//div[@class='dp-value']").text
    # Extract the value of EBITDA
    driver.find_element(by=By.XPATH, value="(.//a[@id='ctl00_ContentPlaceHolder1_ucNavigation_rptNavigation_ctl02_lnkTab'])").click()
    time.sleep(10)
    ebitda = driver.find_element(by=By.XPATH, value="(.//table[@class='sal-summary-section__table']/tbody/tr[4]/td[4])").text
    # Extract the value of fair value
    driver.find_element(by=By.XPATH, value="(.//a[@id='ctl00_ContentPlaceHolder1_ucNavigation_rptNavigation_ctl03_lnkTab'])").click()
    time.sleep(10)
    if(driver.find_element(by=By.XPATH, value="(.//div[@class='sal-columns sal-small-8 sal-medium-12 legend-items']/div)[1]//span").text == 'Fair Value'):
        fair_value = driver.find_element(by=By.XPATH, value="(.//div[@class='sal-columns sal-small-8 sal-medium-12 legend-items']/div)[1]//div[@class='legend-price']/span").text
    else:
        fair_value = driver.find_element(by=By.XPATH, value="(.//div[@class='sal-columns sal-small-8 sal-medium-12 legend-items']/div)[2]//div[@class='legend-price']/span").text
    # Extract the value of total yield of TTM
    total_yield_of_ttm = driver.find_element(by=By.XPATH, value="(.//table[@class='mds-table__sal mds-table--fixed-column__sal']/tbody/tr[5]/td[12])").text



    # write to excel
    xfile = openpyxl.load_workbook(filename=_file_name, read_only=False, keep_vba=True)
    sheet = xfile[_sheet_name]

    # write fair value
    sheet.cell(row=_row, column=16).font = Font(size=14)
    sheet.cell(row=_row, column=16).alignment = Alignment(vertical='center')
    sheet.cell(row=_row, column=16).value = fair_value
    # write total yield of TTM
    sheet.cell(row=_row, column=17).font = Font(size=14)
    sheet.cell(row=_row, column=17).alignment = Alignment(vertical='center')
    sheet.cell(row=_row, column=17).value = total_yield_of_ttm
    # write rating
    sheet.cell(row=_row, column=18).font = Font(size=14)
    sheet.cell(row=_row, column=18).alignment = Alignment(vertical='center')
    sheet.cell(row=_row, column=18).value = rating
    # write price/book
    sheet.cell(row=_row, column=19).font = Font(size=14)
    sheet.cell(row=_row, column=19).alignment = Alignment(vertical='center')
    sheet.cell(row=_row, column=19).value = price_pre_book
    # write price/cash flow
    sheet.cell(row=_row, column=20).font = Font(size=14)
    sheet.cell(row=_row, column=20).alignment = Alignment(vertical='center')
    sheet.cell(row=_row, column=20).value = price_pre_cash_flow
    # write price/sales
    sheet.cell(row=_row, column=21).font = Font(size=14)
    sheet.cell(row=_row, column=21).alignment = Alignment(vertical='center')
    sheet.cell(row=_row, column=21).value = price_pre_sales
    # write price/earnings
    sheet.cell(row=_row, column=22).font = Font(size=14)
    sheet.cell(row=_row, column=22).alignment = Alignment(vertical='center')
    sheet.cell(row=_row, column=22).value = price_pre_earnings
    # write EABDTA
    sheet.cell(row=_row, column=23).font = Font(size=14)
    sheet.cell(row=_row, column=23).alignment = Alignment(vertical='center')
    sheet.cell(row=_row, column=23).value = ebitda

    xfile.save(_file_name)

    _row += 1
    lines[2] = str(_row)
    with open("parameters.txt", "w") as file:
        file.writelines(lines)


    print('rating-> ' + rating)
    print('price_pre_book-> '+ price_pre_book)
    print('price_pre_cash_flow-> '+ price_pre_cash_flow)
    print('price_pre_sales-> '+ price_pre_sales)
    print('price_pre_earnings-> '+ price_pre_earnings)
    print('ebitda-> '+ ebitda)
    print('fair_value-> '+ fair_value)
    print('total_yield_of_ttm-> '+ total_yield_of_ttm)
    print('\n\n\n\n')


